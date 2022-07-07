import openpyxl

# 公式
workbookCal = openpyxl.Workbook()
sheet = workbookCal.active
sheet['A1'] = 200
sheet['A2'] = 300
# 设置公式.
sheet['A3'] = '=SUM(A1:A2)'
workbookCal.save('.\\source\\writeFormula.xlsx')

# 行、列操作
workbookOpt = openpyxl.Workbook()
sheetOpt = workbookOpt.active
sheetOpt['A1'] = 'Tall row'
sheetOpt['B2'] = 'Wide column'
# 设置宽高
sheetOpt.row_dimensions[1].height = 70
sheetOpt.column_dimensions['B'].width = 20
# 合并拆分单元格
sheetOpt.merge_cells('A1:D3')
sheetOpt['A1'] = 'Twelve cells merged together.'
sheetOpt.merge_cells('C5:D5')
sheetOpt['C5'] = 'Two merged cells.'
workbookOpt.save('.\\source\\dimensions.xlsx')
# 分拆单元格
sheetOpt.unmerge_cells('C5:D5')
workbookOpt.save('.\\source\\dimensions.xlsx')
# 冻结窗口
sheetOpt.freeze_panes = 'C5'
workbookOpt.save('.\\source\\dimensions.xlsx')

#图标
workbookDraw = openpyxl.Workbook()
sheetDraw = workbookDraw.active
for i in range(1, 11):
    sheetDraw['A' + str(i)] = i
refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1,max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.BarChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheetDraw.add_chart(chartObj, 'C5')
workbookDraw.save('.\\source\\sampleChart.xlsx')