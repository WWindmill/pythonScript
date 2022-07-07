import openpyxl

# 工作簿对象
workbook = openpyxl.load_workbook(".\\source\\example.xlsx")
print('type of result: ', type(workbook))
print("all sheet names: ", workbook.sheetnames)
# 获取工作表
sheet = workbook['Sheet1']
print("sheet obj: ", sheet, " sheet title:",  sheet.title)
# 修改sheet名称 并转储为另一个文件
sheet.title = 'Spam Spam Spam'
workbook.save('.\\source\\example_copy.xlsx')
# 创建和删除工作表
workbook.create_sheet(index=3, title="the fourth sheet")
print('sheet names: ', workbook.sheetnames)
del workbook['the fourth sheet']
print('sheet names: ', workbook.sheetnames)
workbook.save('.\\source\\example_copy.xlsx')
# 修改单元格属性值
sheet['B1'] = 'Hello, world!'
print('B1 modified value: ', sheet['B1'].value)
workbook.save('.\\source\\example_copy.xlsx')