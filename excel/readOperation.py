import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 工作簿对象
workbook = openpyxl.load_workbook("C:\\projects\\Python\\9.源代码文件\\automate_online-materials\\example.xlsx")
print('type of result: ', type(workbook))
print("all sheet names: ", workbook.sheetnames)
# 获取工作表
sheet = workbook['Sheet1']
print("sheet obj: ", sheet, " sheet title:",  sheet.title)
# 获取工作簿的活动表
anotherSheet = workbook.active
print("active sheet: ", anotherSheet)
# 获取单元格
print("cell A1: ", sheet['A1'])
print("cell A1 val: ", sheet['A1'].value)
c = sheet['B1']
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
print('Cell %s is %s' % (c.coordinate, c.value))
print("cell[B1]: ", sheet.cell(row=1, column=2))
# 步长为2
for i in range(1, 8, 2):
    print('row:%s,column:2, value:%s' % (i, sheet.cell(row=i, column=2).value))
# 获取工作表大小
print("max row: ", sheet.max_row)
print("max column: ", sheet.max_column)
# 列字母与数字转换
print("1 mean letter: ", get_column_letter(1))
print(sheet.max_column, "mean letter: ", get_column_letter(sheet.max_column))
print("column A point at num: ", column_index_from_string('A'))
# 按行遍历 method1
print(tuple(sheet['A1':'C3']))
for rowCell in sheet['A1':'C3']:
    for eachCell in rowCell:
        print(eachCell.coordinate, eachCell.value)
    print('--- END OF ROW ---')
# 按行遍历 method2
print(list(sheet.rows)[0])
for cellObj in list(sheet.rows)[0]:
 print(cellObj.value)
print('--- END OF ROW ---')
# 按列遍历
print(list(sheet.columns)[0])
for cellObj in list(sheet.columns)[0]:
 print(cellObj.value)
print('--- END OF column ---')